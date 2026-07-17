"""Tests for multi-sheet emotion prediction behavior."""

from __future__ import annotations

import tempfile
import unittest
import json
from datetime import datetime
from pathlib import Path
import sys
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

import predict_emotion
from profile_loader import load_profile_registry


class MultiSheetPredictionTests(unittest.TestCase):
    def test_all_sheets_switch_profile_backgrounds(self) -> None:
        captured_backgrounds: list[str] = []

        def fake_predict_row(**kwargs):
            captured_backgrounds.append(kwargs["prompt_parts"]["case_background"])
            return "neutral", "neutral", "原因", "原始输出"

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_path = root / "dialogue.xlsx"
            output_path = root / "output.xlsx"
            workbook = Workbook()
            workbook.active.title = "index"
            for sheet_name in ("one", "two"):
                sheet = workbook.create_sheet(sheet_name)
                sheet.append(["轮次", "提问人员", "被谈话人"])
                sheet.append([1, "问题", "回答"])
            workbook.save(input_path)

            (root / "one.txt").write_text("人物一完整背景", encoding="utf-8")
            (root / "two.txt").write_text("人物二完整背景", encoding="utf-8")
            config_path = root / "profiles.json"
            config_path.write_text(
                json.dumps(
                    {
                        "profiles": {
                            "p1": {"respondent_name": "人物一", "story_background": "one.txt"},
                            "p2": {"respondent_name": "人物二", "story_background": "two.txt"},
                        },
                        "bindings": {
                            "dialogue.xlsx::one": "p1",
                            "dialogue.xlsx::two": "p2",
                        },
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            with patch.object(predict_emotion, "predict_row", side_effect=fake_predict_row):
                predict_emotion.process_all_sheets(
                    input_path=input_path,
                    output_path=output_path,
                    client=object(),
                    model="test-model",
                    prompt_parts={"person_profile": "旧人物", "case_background": "旧背景"},
                    temperature=0.0,
                    max_tokens=10,
                    history_turns=None,
                    limit=None,
                    profile_registry=load_profile_registry(config_path),
                    allow_legacy_profile=False,
                )

        self.assertEqual(captured_backgrounds, ["人物一完整背景", "人物二完整背景"])

    def test_all_sheets_skip_index_and_reset_history(self) -> None:
        captured_histories: list[str] = []

        def fake_predict_row(**kwargs):
            captured_histories.append(kwargs["dialogue_history"])
            return "neutral", "neutral", "测试原因", "Emotion：neutral\nReason：测试原因"

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_path = root / "input.xlsx"
            output_path = root / "output.xlsx"
            workbook = Workbook()
            index = workbook.active
            index.title = "index"
            index.append(["sheet_name", "chat_id"])
            index.append(["chat-1", "chat-1"])

            for sheet_name, prefix in (("chat-1", "甲"), ("chat-2", "乙")):
                sheet = workbook.create_sheet(sheet_name)
                sheet.append(["轮次", "提问人员", "被谈话人"])
                sheet.append([1, f"{prefix}问题1", f"{prefix}回答1"])
                sheet.append([2, f"{prefix}问题2", f"{prefix}回答2"])
            workbook.save(input_path)

            with patch.object(predict_emotion, "predict_row", side_effect=fake_predict_row):
                processed = predict_emotion.process_all_sheets(
                    input_path=input_path,
                    output_path=output_path,
                    client=object(),
                    model="test-model",
                    prompt_parts={},
                    temperature=0.0,
                    max_tokens=10,
                    history_turns=None,
                    limit=None,
                    sheet_limit=None,
                )

            self.assertEqual(4, processed)
            self.assertEqual("", captured_histories[0])
            self.assertIn("甲问题1", captured_histories[1])
            self.assertEqual("", captured_histories[2])
            self.assertNotIn("甲", captured_histories[2])

            result = load_workbook(output_path)
            self.assertEqual(["index", "chat-1", "chat-2"], result.sheetnames)
            self.assertEqual("预测情绪标签", result["chat-1"].cell(1, 4).value)
            self.assertEqual("neutral", result["chat-2"].cell(2, 4).value)
            self.assertEqual(2, result["index"].max_column)

    def test_all_sheets_limit_is_global(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_path = root / "input.xlsx"
            output_path = root / "output.xlsx"
            workbook = Workbook()
            workbook.active.title = "index"
            for sheet_name in ("one", "two"):
                sheet = workbook.create_sheet(sheet_name)
                sheet.append(["轮次", "提问人员", "被谈话人"])
                sheet.append([1, "问题", "回答"])
                sheet.append([2, "问题", "回答"])
            workbook.save(input_path)

            with patch.object(
                predict_emotion,
                "predict_row",
                return_value=("neutral", "neutral", "原因", "原始输出"),
            ):
                processed = predict_emotion.process_all_sheets(
                    input_path,
                    output_path,
                    object(),
                    "test-model",
                    {},
                    0.0,
                    10,
                    None,
                    limit=3,
                    sheet_limit=None,
                )

            self.assertEqual(3, processed)
            result = load_workbook(output_path)
            self.assertEqual("neutral", result["one"].cell(3, 4).value)
            self.assertEqual("neutral", result["two"].cell(2, 4).value)
            self.assertIsNone(result["two"].cell(3, 4).value)

    def test_sheet_limit_processes_complete_selected_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_path = root / "input.xlsx"
            output_path = root / "output.xlsx"
            workbook = Workbook()
            workbook.active.title = "index"
            for sheet_name in ("one", "two", "three"):
                sheet = workbook.create_sheet(sheet_name)
                sheet.append(["轮次", "提问人员", "被谈话人"])
                sheet.append([1, f"{sheet_name}问题1", "回答1"])
                sheet.append([2, f"{sheet_name}问题2", "回答2"])
            workbook.save(input_path)

            with patch.object(
                predict_emotion,
                "predict_row",
                return_value=("neutral", "neutral", "原因", "原始输出"),
            ):
                processed = predict_emotion.process_all_sheets(
                    input_path,
                    output_path,
                    object(),
                    "test-model",
                    {},
                    0.0,
                    10,
                    history_turns=None,
                    limit=None,
                    sheet_limit=2,
                )

            self.assertEqual(4, processed)
            result = load_workbook(output_path)
            self.assertEqual("neutral", result["one"].cell(3, 4).value)
            self.assertEqual("neutral", result["two"].cell(3, 4).value)
            self.assertIsNone(result["three"].cell(1, 4).value)

    def test_history_window_keeps_only_recent_turns(self) -> None:
        import pandas as pd

        df = pd.DataFrame(
            {
                "提问人员": ["问题1", "问题2", "问题3", "问题4"],
                "被谈话人": ["回答1", "回答2", "回答3", "回答4"],
            }
        )
        history, current_question = predict_emotion.get_prompt_inputs(
            df, 3, "提问人员", "被谈话人", history_turns=2
        )
        self.assertEqual("问题4", current_question)
        self.assertNotIn("问题1", history)
        self.assertIn("问题2", history)
        self.assertIn("问题3", history)
        self.assertNotIn("回答4", history)

    def test_output_path_gets_timestamp(self) -> None:
        path = predict_emotion.add_timestamp_to_path(
            Path("outputs/result.xlsx"), datetime(2026, 7, 13, 14, 5, 6)
        )
        self.assertEqual(Path("outputs/result_20260713_140506.xlsx"), path)


if __name__ == "__main__":
    unittest.main()
