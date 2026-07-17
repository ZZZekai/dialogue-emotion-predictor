import argparse
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

from profile_loader import (
    ProfileRegistry,
    apply_profile_to_emotion_prompt_parts,
    load_profile_registry,
)

from utils import (
    ALLOWED_EMOTIONS,
    build_prompt,
    call_model_with_retry,
    create_openai_client,
    ensure_parent_dir,
    get_model_name,
    is_blank,
    load_prompt_parts,
    parse_emotion_response,
    postprocess_emotion,
    to_text,
)


# Column names are centralized here so they can be changed for new Excel formats.
HISTORY_COL = "对话历史"
QUESTION_COL = "审调人员当前问题"
FALLBACK_QUESTION_COL = "审调人员问题"
FALLBACK_ANSWER_COL = "被谈话人回答"

# Extra aliases keep compatibility with common transcript exports.
QUESTION_COL_CANDIDATES = [QUESTION_COL, FALLBACK_QUESTION_COL, "提问人员"]
ANSWER_COL_CANDIDATES = [FALLBACK_ANSWER_COL, "被谈话人"]

OUTPUT_EMOTION_COL = "预测情绪标签"
OUTPUT_RAW_EMOTION_COL = "原始情绪标签"
OUTPUT_REASON_COL = "预测原因"
OUTPUT_RAW_MODEL_COL = "原始模型输出"

TEMP_OUTPUT_PATH = Path("outputs/temp_result.xlsx")
AUTOSAVE_EVERY = 20
MAX_RETRIES = 3
RETRY_WAIT_SECONDS = 2
INDEX_SHEET_NAME = "index"

LOGGER = logging.getLogger("emotion_predictor")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Predict respondent emotions from interrogation dialogue Excel data."
    )
    parser.add_argument("--input", required=True, help="Input Excel path.")
    parser.add_argument("--output", required=True, help="Output Excel path.")
    parser.add_argument("--model", default=None, help="DashScope model name, e.g. qwen-plus.")
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Maximum predictions: rows in single-sheet mode, total rows in all-sheets mode.",
    )
    parser.add_argument(
        "--history-turns",
        type=int,
        default=None,
        help="Keep only the most recent N dialogue turns. Default: full history.",
    )
    parser.add_argument(
        "--sheet-limit",
        type=int,
        default=None,
        help="In all-sheets mode, process only the first N dialogue worksheets.",
    )
    parser.add_argument(
        "--no-timestamp",
        action="store_true",
        help="Use the exact --output path instead of appending a timestamp.",
    )
    sheet_mode = parser.add_mutually_exclusive_group()
    sheet_mode.add_argument(
        "--sheet", default=0, help="Excel sheet name or index. Default: first sheet."
    )
    sheet_mode.add_argument(
        "--all-sheets",
        action="store_true",
        help="Process every dialogue sheet independently and skip the index sheet.",
    )
    parser.add_argument("--env", default=".env", help="Environment file path. Default: .env.")
    parser.add_argument("--prompts-dir", default="prompts", help="Prompt directory. Default: prompts.")
    parser.add_argument(
        "--profile-config",
        default="local_profiles/profiles.json",
        help="Shared local profile registry JSON.",
    )
    parser.add_argument(
        "--profile-id",
        default=None,
        help="Explicit profile override. Otherwise resolve by input filename and sheet.",
    )
    parser.add_argument(
        "--allow-legacy-profile",
        action="store_true",
        help="Allow unmapped sheets to fall back to prompts/person_profile.txt and case_background.txt.",
    )
    parser.add_argument("--temperature", type=float, default=0.0, help="Model temperature.")
    parser.add_argument("--max-tokens", type=int, default=512, help="Max output tokens.")
    return parser.parse_args()


def parse_sheet_arg(sheet: str) -> str | int:
    return int(sheet) if str(sheet).isdigit() else sheet


def detect_column(df: pd.DataFrame, candidates: list[str], column_role: str) -> str:
    for column in candidates:
        if column in df.columns:
            return column
    raise ValueError(
        f"Missing {column_role} column. Supported column names: {', '.join(candidates)}"
    )


def detect_optional_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    return next((column for column in candidates if column in df.columns), None)


def validate_input_columns(df: pd.DataFrame) -> tuple[str, str | None]:
    question_col = detect_column(df, QUESTION_COL_CANDIDATES, "question")
    answer_col = detect_optional_column(df, ANSWER_COL_CANDIDATES)

    if HISTORY_COL not in df.columns and not answer_col:
        raise ValueError(
            f"Input Excel must contain '{HISTORY_COL}', or contain a respondent answer column "
            f"to construct history automatically. Supported answer columns: "
            f"{', '.join(ANSWER_COL_CANDIDATES)}"
        )

    return question_col, answer_col


def build_history_from_previous_rows(
    df: pd.DataFrame,
    row_position: int,
    question_col: str,
    answer_col: str,
    history_turns: int | None = None,
) -> str:
    # Important: range(..., row_position) excludes the current row, so the current
    # respondent answer can never enter dialogue_history in auto-history mode.
    start_position = (
        0 if history_turns is None else max(0, row_position - history_turns)
    )
    lines: list[str] = []
    for previous_position in range(start_position, row_position):
        previous_row = df.iloc[previous_position]
        previous_question = to_text(previous_row.get(question_col, ""))
        previous_answer = to_text(previous_row.get(answer_col, ""))

        if previous_question:
            lines.append(f"审调人员：{previous_question}")
        if previous_answer:
            lines.append(f"被谈话人：{previous_answer}")

    return "\n".join(lines)


def get_prompt_inputs(
    df: pd.DataFrame,
    row_position: int,
    question_col: str,
    answer_col: str | None,
    history_turns: int | None = None,
) -> tuple[str, str]:
    row = df.iloc[row_position]
    current_question = to_text(row.get(question_col, ""))

    if HISTORY_COL in df.columns:
        dialogue_history = to_text(row.get(HISTORY_COL, ""))
        if history_turns is not None:
            dialogue_history = limit_provided_history(dialogue_history, history_turns)
        ensure_current_answer_not_in_history(
            df=df,
            row_position=row_position,
            answer_col=answer_col,
            dialogue_history=dialogue_history,
        )
        return dialogue_history, current_question

    if not answer_col:
        raise ValueError(f"Cannot build dialogue history without an answer column.")

    dialogue_history = build_history_from_previous_rows(
        df=df,
        row_position=row_position,
        question_col=question_col,
        answer_col=answer_col,
        history_turns=history_turns,
    )
    return dialogue_history, current_question


def limit_provided_history(dialogue_history: str, history_turns: int) -> str:
    """Keep the last N question-led blocks from a prebuilt history string."""
    if history_turns <= 0 or not dialogue_history:
        return ""
    question_marker = "审调人员："
    normalized = dialogue_history.replace("审调人员:", question_marker)
    positions = [
        position
        for position in range(len(normalized))
        if normalized.startswith(question_marker, position)
    ]
    if len(positions) <= history_turns:
        return normalized
    return normalized[positions[-history_turns] :].lstrip()


def ensure_current_answer_not_in_history(
    df: pd.DataFrame,
    row_position: int,
    answer_col: str | None,
    dialogue_history: str,
) -> None:
    if not answer_col:
        return

    current_answer = to_text(df.iloc[row_position].get(answer_col, ""))
    if current_answer and current_answer in dialogue_history:
        row_label = df.index[row_position]
        raise ValueError(
            f"Data leakage detected at row {row_label}: current respondent answer appears "
            f"in '{HISTORY_COL}'. Remove the current answer from history before prediction."
        )


def should_predict_row(df: pd.DataFrame, row_position: int, answer_col: str | None) -> bool:
    # If an answer column exists, blank respondent-answer rows are usually interviewer-only
    # fragments and should not be predicted as respondent emotions.
    if not answer_col:
        return True
    return not is_blank(df.iloc[row_position].get(answer_col, ""))


def predict_row(
    client,
    model: str,
    prompt_parts: dict[str, str],
    dialogue_history: str,
    current_question: str,
    temperature: float,
    max_tokens: int,
) -> tuple[str, str, str, str]:
    prompt = build_prompt(
        prompt_template=prompt_parts["prompt_template"],
        person_profile=prompt_parts["person_profile"],
        case_background=prompt_parts["case_background"],
        emotion_examples=prompt_parts["emotion_examples"],
        dialogue_history=dialogue_history,
        current_question=current_question,
    )

    raw_output = call_model_with_retry(
        client=client,
        model=model,
        prompt=prompt,
        max_retries=MAX_RETRIES,
        retry_wait_seconds=RETRY_WAIT_SECONDS,
        temperature=temperature,
        max_tokens=max_tokens,
    )
    raw_emotion, reason = parse_emotion_response(raw_output, ALLOWED_EMOTIONS)
    final_emotion = (
        "PARSE_ERROR"
        if raw_emotion == "PARSE_ERROR"
        else postprocess_emotion(raw_emotion, reason)
    )
    return final_emotion, raw_emotion, reason, raw_output


def save_excel(df: pd.DataFrame, output_path: Path) -> None:
    ensure_parent_dir(output_path)
    df.to_excel(output_path, index=False)


def initialize_output_columns(df: pd.DataFrame) -> None:
    """Add result columns without overwriting existing predictions."""
    for output_col in [
        OUTPUT_EMOTION_COL,
        OUTPUT_RAW_EMOTION_COL,
        OUTPUT_REASON_COL,
        OUTPUT_RAW_MODEL_COL,
    ]:
        if output_col not in df.columns:
            df[output_col] = ""


def process_dataframe(
    df: pd.DataFrame,
    client,
    model: str,
    prompt_parts: dict[str, str],
    temperature: float,
    max_tokens: int,
    history_turns: int | None = None,
    limit: int | None = None,
    progress: tqdm | None = None,
) -> tuple[pd.DataFrame, int]:
    """Predict one isolated dialogue table and return its processed row count."""
    work_df = df.copy()
    question_col, answer_col = validate_input_columns(work_df)
    initialize_output_columns(work_df)

    row_positions = [
        row_position
        for row_position in range(len(work_df))
        if should_predict_row(work_df, row_position, answer_col)
    ]
    if limit is not None:
        row_positions = row_positions[:limit]

    local_progress = progress or tqdm(
        total=len(row_positions), desc="Predicting emotions"
    )
    owns_progress = progress is None

    try:
        for row_position in row_positions:
            try:
                dialogue_history, current_question = get_prompt_inputs(
                    df=work_df,
                    row_position=row_position,
                    question_col=question_col,
                    answer_col=answer_col,
                    history_turns=history_turns,
                )
                final_emotion, raw_emotion, reason, raw_output = predict_row(
                    client=client,
                    model=model,
                    prompt_parts=prompt_parts,
                    dialogue_history=dialogue_history,
                    current_question=current_question,
                    temperature=temperature,
                    max_tokens=max_tokens,
                )
            except Exception as exc:
                final_emotion = "API_ERROR"
                raw_emotion = "API_ERROR"
                reason = str(exc)
                raw_output = str(exc)

            row_index = work_df.index[row_position]
            work_df.at[row_index, OUTPUT_EMOTION_COL] = final_emotion
            work_df.at[row_index, OUTPUT_RAW_EMOTION_COL] = raw_emotion
            work_df.at[row_index, OUTPUT_REASON_COL] = reason
            work_df.at[row_index, OUTPUT_RAW_MODEL_COL] = raw_output
            local_progress.update(1)
    finally:
        if owns_progress:
            local_progress.close()

    return work_df, len(row_positions)


def worksheet_to_dataframe(worksheet) -> pd.DataFrame:
    """Convert worksheet values to a DataFrame while retaining worksheet row order."""
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    return pd.DataFrame(rows[1:], columns=headers)


def write_dataframe_results_to_worksheet(df: pd.DataFrame, worksheet) -> None:
    """Write only prediction columns, preserving the source workbook structure."""
    header_map = {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }
    next_column = worksheet.max_column + 1
    for output_col in [
        OUTPUT_EMOTION_COL,
        OUTPUT_RAW_EMOTION_COL,
        OUTPUT_REASON_COL,
        OUTPUT_RAW_MODEL_COL,
    ]:
        column = header_map.get(output_col)
        if column is None:
            column = next_column
            next_column += 1
            worksheet.cell(1, column, output_col)
            header_map[output_col] = column
        for row_position, value in enumerate(df[output_col], start=2):
            worksheet.cell(row_position, column, "" if is_blank(value) else value)


def process_all_sheets(
    input_path: Path,
    output_path: Path,
    client,
    model: str,
    prompt_parts: dict[str, str],
    temperature: float,
    max_tokens: int,
    history_turns: int | None,
    limit: int | None,
    sheet_limit: int | None = None,
    profile_registry: ProfileRegistry | None = None,
    profile_id: str | None = None,
    allow_legacy_profile: bool = True,
) -> int:
    """Process every chat sheet with history reset at each worksheet boundary."""
    workbook = load_workbook(input_path)
    dialogue_sheet_names = [
        name for name in workbook.sheetnames if name.casefold() != INDEX_SHEET_NAME
    ]
    if sheet_limit is not None:
        dialogue_sheet_names = dialogue_sheet_names[:sheet_limit]

    eligible_counts: dict[str, int] = {}
    for sheet_name in dialogue_sheet_names:
        df = worksheet_to_dataframe(workbook[sheet_name])
        if df.empty:
            eligible_counts[sheet_name] = 0
            continue
        try:
            _, answer_col = validate_input_columns(df)
            eligible_counts[sheet_name] = sum(
                should_predict_row(df, row_position, answer_col)
                for row_position in range(len(df))
            )
        except ValueError as exc:
            LOGGER.warning("跳过工作表 %s：%s", sheet_name, exc)
            eligible_counts[sheet_name] = 0

    total_eligible = sum(eligible_counts.values())
    total_target = min(total_eligible, limit) if limit is not None else total_eligible
    processed_total = 0
    processed_sheets = 0
    temp_path = output_path.with_name(f"{output_path.stem}_temp.xlsx")

    with tqdm(total=total_target, desc="Predicting all sheets") as progress:
        for sheet_name in dialogue_sheet_names:
            if limit is not None and processed_total >= limit:
                break
            if eligible_counts[sheet_name] == 0:
                continue

            worksheet = workbook[sheet_name]
            df = worksheet_to_dataframe(worksheet)
            remaining = None if limit is None else limit - processed_total
            sheet_prompt_parts = prompt_parts
            if profile_registry is not None:
                try:
                    profile = profile_registry.resolve(
                        input_path, sheet_name, profile_id=profile_id
                    )
                    sheet_prompt_parts = apply_profile_to_emotion_prompt_parts(
                        prompt_parts, profile
                    )
                    LOGGER.info(
                        "工作表 %s 使用人物档案 %s（%s）",
                        sheet_name,
                        profile.profile_id,
                        profile.respondent_name,
                    )
                except KeyError:
                    if not allow_legacy_profile:
                        raise
                    LOGGER.warning("工作表 %s 未绑定档案，使用旧版 Prompt 背景", sheet_name)
            result_df, processed = process_dataframe(
                df=df,
                client=client,
                model=model,
                prompt_parts=sheet_prompt_parts,
                temperature=temperature,
                max_tokens=max_tokens,
                history_turns=history_turns,
                limit=remaining,
                progress=progress,
            )
            write_dataframe_results_to_worksheet(result_df, worksheet)
            processed_total += processed
            processed_sheets += 1

            # Saving every row would repeatedly rewrite a very large workbook.
            if processed_sheets % AUTOSAVE_EVERY == 0:
                ensure_parent_dir(temp_path)
                workbook.save(temp_path)
                LOGGER.info("已保存多表临时结果：%s", temp_path)

    ensure_parent_dir(output_path)
    workbook.save(output_path)
    return processed_total


def add_timestamp_to_path(path: Path, now: datetime | None = None) -> Path:
    """Append a sortable local timestamp before the file extension."""
    timestamp = (now or datetime.now()).strftime("%Y%m%d_%H%M%S")
    return path.with_name(f"{path.stem}_{timestamp}{path.suffix}")


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    if not args.no_timestamp:
        output_path = add_timestamp_to_path(output_path)

    if args.history_turns is not None and args.history_turns < 0:
        raise ValueError("--history-turns must be 0 or a positive integer.")
    if args.sheet_limit is not None and args.sheet_limit <= 0:
        raise ValueError("--sheet-limit must be a positive integer.")
    if args.sheet_limit is not None and not args.all_sheets:
        raise ValueError("--sheet-limit can only be used together with --all-sheets.")

    if not input_path.exists():
        raise FileNotFoundError(f"Input Excel file not found: {input_path}")

    prompt_parts = load_prompt_parts(args.prompts_dir)
    try:
        profile_registry = load_profile_registry(args.profile_config)
    except FileNotFoundError:
        if not args.allow_legacy_profile:
            raise
        profile_registry = None
        LOGGER.warning("未找到人物档案注册表，使用旧版 Prompt 背景")
    client = create_openai_client(args.env)
    model = get_model_name(args.model)

    if args.all_sheets:
        processed_total = process_all_sheets(
            input_path=input_path,
            output_path=output_path,
            client=client,
            model=model,
            prompt_parts=prompt_parts,
            temperature=args.temperature,
            max_tokens=args.max_tokens,
            history_turns=args.history_turns,
            limit=args.limit,
            sheet_limit=args.sheet_limit,
            profile_registry=profile_registry,
            profile_id=args.profile_id,
            allow_legacy_profile=args.allow_legacy_profile,
        )
        print(f"Saved result to: {output_path}")
        print(f"Processed prediction rows: {processed_total}")
        return

    selected_sheet = parse_sheet_arg(args.sheet)
    excel_file = pd.ExcelFile(input_path)
    sheet_name = (
        excel_file.sheet_names[selected_sheet]
        if isinstance(selected_sheet, int)
        else selected_sheet
    )
    sheet_prompt_parts = prompt_parts
    if profile_registry is not None:
        try:
            profile = profile_registry.resolve(
                input_path, sheet_name, profile_id=args.profile_id
            )
            sheet_prompt_parts = apply_profile_to_emotion_prompt_parts(
                prompt_parts, profile
            )
            LOGGER.info(
                "工作表 %s 使用人物档案 %s（%s）",
                sheet_name,
                profile.profile_id,
                profile.respondent_name,
            )
        except KeyError:
            if not args.allow_legacy_profile:
                raise
            LOGGER.warning("工作表 %s 未绑定档案，使用旧版 Prompt 背景", sheet_name)

    df = pd.read_excel(input_path, sheet_name=sheet_name)
    work_df, processed_count = process_dataframe(
        df=df,
        client=client,
        model=model,
        prompt_parts=sheet_prompt_parts,
        temperature=args.temperature,
        max_tokens=args.max_tokens,
        history_turns=args.history_turns,
        limit=args.limit,
    )

    if processed_count >= AUTOSAVE_EVERY:
        save_excel(work_df, TEMP_OUTPUT_PATH)

    save_excel(work_df, output_path)
    print(f"Saved result to: {output_path}")


if __name__ == "__main__":
    main()
