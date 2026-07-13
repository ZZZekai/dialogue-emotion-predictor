"""Merge prediction columns from two emotion experiment workbooks."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


TURN_COL = "轮次"
QUESTION_COL = "提问人员"
ANSWER_COL = "被谈话人"
INDEX_SHEET = "index"

PREDICTION_FIELDS = (
    "预测情绪标签",
    "原始情绪标签",
    "预测原因",
    "原始模型输出",
)
COMPARISON_HEADERS = tuple(
    f"{field_name}{experiment}"
    for field_name in PREDICTION_FIELDS
    for experiment in (1, 2)
)


@dataclass
class MergeReport:
    """Human-readable merge statistics for the GUI."""

    compared_sheets: int = 0
    compared_turns: int = 0
    different_emotions: int = 0
    missing_sheets_in_file2: list[str] = field(default_factory=list)
    missing_turns_in_file2: int = 0
    dialogue_mismatches: int = 0
    skipped_sheets: list[str] = field(default_factory=list)

    def summary(self) -> str:
        return (
            f"已比较工作表：{self.compared_sheets}\n"
            f"已匹配轮次：{self.compared_turns}\n"
            f"情绪标签不同：{self.different_emotions}\n"
            f"文件2缺少工作表：{len(self.missing_sheets_in_file2)}\n"
            f"文件2缺少轮次：{self.missing_turns_in_file2}\n"
            f"对话文本不一致：{self.dialogue_mismatches}\n"
            f"跳过的工作表：{len(self.skipped_sheets)}"
        )


def _normalize_turn(value: Any) -> str:
    """Normalize numeric and textual turn identifiers for matching."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_map(worksheet: Worksheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None and str(cell.value).strip()
    }


def _validate_dialogue_sheet(worksheet: Worksheet) -> dict[str, int]:
    headers = _header_map(worksheet)
    missing = [name for name in (TURN_COL, QUESTION_COL, ANSWER_COL) if name not in headers]
    if missing:
        raise ValueError(f"工作表 {worksheet.title!r} 缺少列：{', '.join(missing)}")
    return headers


def _rows_by_turn(
    worksheet: Worksheet,
    headers: dict[str, int],
) -> tuple[dict[str, dict[str, Any]], int]:
    """Collect dialogue and prediction values keyed by turn number."""
    rows: dict[str, dict[str, Any]] = {}
    duplicates = 0
    for row_number in range(2, worksheet.max_row + 1):
        turn = _normalize_turn(worksheet.cell(row_number, headers[TURN_COL]).value)
        if not turn:
            continue
        if turn in rows:
            duplicates += 1
            continue
        rows[turn] = {
            "row_number": row_number,
            QUESTION_COL: worksheet.cell(row_number, headers[QUESTION_COL]).value,
            ANSWER_COL: worksheet.cell(row_number, headers[ANSWER_COL]).value,
            **{
                field_name: (
                    worksheet.cell(row_number, headers[field_name]).value
                    if field_name in headers
                    else None
                )
                for field_name in PREDICTION_FIELDS
            },
        }
    return rows, duplicates


def _remove_old_prediction_columns(worksheet: Worksheet) -> None:
    headers = _header_map(worksheet)
    removable = set(PREDICTION_FIELDS) | set(COMPARISON_HEADERS)
    columns = sorted(
        (column for name, column in headers.items() if name in removable), reverse=True
    )
    for column in columns:
        worksheet.delete_cols(column)


def _copy_header_style(source_cell, target_cell) -> None:
    if source_cell is None:
        return
    target_cell._style = copy(source_cell._style)
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format


def _append_comparison_columns(
    worksheet: Worksheet,
    source_rows: dict[str, dict[str, Any]],
    comparison_rows: dict[str, dict[str, Any]],
    report: MergeReport,
    header_style_source,
) -> None:
    headers = _header_map(worksheet)
    first_column = worksheet.max_column + 1
    disagreement_fill = PatternFill("solid", fgColor="FFF2CC")

    for offset, header in enumerate(COMPARISON_HEADERS):
        cell = worksheet.cell(1, first_column + offset, header)
        _copy_header_style(header_style_source, cell)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.column_dimensions[cell.column_letter].width = (
            20 if "标签" in header else 55
        )

    result_columns = {
        header: first_column + offset for offset, header in enumerate(COMPARISON_HEADERS)
    }

    for turn, source_data in source_rows.items():
        row_number = int(source_data["row_number"])
        comparison_data = comparison_rows.get(turn)
        if comparison_data is None:
            report.missing_turns_in_file2 += 1

        for field_name in PREDICTION_FIELDS:
            value1 = source_data.get(field_name)
            value2 = comparison_data.get(field_name) if comparison_data else None
            worksheet.cell(row_number, result_columns[f"{field_name}1"], value1)
            worksheet.cell(row_number, result_columns[f"{field_name}2"], value2)

        for column in result_columns.values():
            worksheet.cell(row_number, column).alignment = Alignment(
                vertical="top", wrap_text=True
            )

        if comparison_data is not None:
            report.compared_turns += 1
            question_matches = str(source_data.get(QUESTION_COL) or "").strip() == str(
                comparison_data.get(QUESTION_COL) or ""
            ).strip()
            answer_matches = str(source_data.get(ANSWER_COL) or "").strip() == str(
                comparison_data.get(ANSWER_COL) or ""
            ).strip()
            if not question_matches or not answer_matches:
                report.dialogue_mismatches += 1

            emotion1 = str(source_data.get("预测情绪标签") or "").strip()
            emotion2 = str(comparison_data.get("预测情绪标签") or "").strip()
            if emotion1 != emotion2:
                report.different_emotions += 1
                worksheet.cell(
                    row_number, result_columns["预测情绪标签1"]
                ).fill = disagreement_fill
                worksheet.cell(
                    row_number, result_columns["预测情绪标签2"]
                ).fill = disagreement_fill

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"


def merge_experiment_workbooks(
    file1: str | Path,
    file2: str | Path,
    output_path: str | Path,
) -> MergeReport:
    """Merge two experiment workbooks by identical sheet name and turn number."""
    source_path = Path(file1).expanduser().resolve()
    comparison_path = Path(file2).expanduser().resolve()
    destination = Path(output_path).expanduser().resolve()

    for path in (source_path, comparison_path):
        if not path.is_file():
            raise FileNotFoundError(f"找不到 Excel 文件：{path}")
        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"只支持 .xlsx 文件：{path.name}")
    if source_path == comparison_path:
        raise ValueError("请选择两个不同的 Excel 文件。")
    if destination in (source_path, comparison_path):
        raise ValueError("输出文件不能覆盖任一输入文件。")

    workbook1 = load_workbook(source_path)
    workbook2 = load_workbook(comparison_path, read_only=True, data_only=False)
    report = MergeReport()

    try:
        for sheet_name in workbook1.sheetnames:
            if sheet_name.casefold() == INDEX_SHEET:
                continue
            if sheet_name not in workbook2.sheetnames:
                report.missing_sheets_in_file2.append(sheet_name)
                continue

            worksheet1 = workbook1[sheet_name]
            worksheet2 = workbook2[sheet_name]
            try:
                headers1 = _validate_dialogue_sheet(worksheet1)
                headers2 = _validate_dialogue_sheet(worksheet2)
            except ValueError:
                report.skipped_sheets.append(sheet_name)
                continue

            has_predictions = any(
                field in headers1 or field in headers2 for field in PREDICTION_FIELDS
            )
            if not has_predictions:
                continue

            rows1, duplicates1 = _rows_by_turn(worksheet1, headers1)
            rows2, duplicates2 = _rows_by_turn(worksheet2, headers2)
            report.dialogue_mismatches += duplicates1 + duplicates2

            style_column = next(
                (headers1[field] for field in PREDICTION_FIELDS if field in headers1),
                headers1[ANSWER_COL],
            )
            header_style_source = copy(worksheet1.cell(1, style_column))
            _remove_old_prediction_columns(worksheet1)
            _append_comparison_columns(
                worksheet1,
                rows1,
                rows2,
                report,
                header_style_source,
            )
            report.compared_sheets += 1

        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook1.save(destination)
    finally:
        workbook2.close()
        workbook1.close()

    return report
