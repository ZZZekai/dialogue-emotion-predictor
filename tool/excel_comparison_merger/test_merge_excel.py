"""Tests for the experiment workbook merger."""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

from merge_excel import COMPARISON_HEADERS, merge_experiment_workbooks


def create_experiment(path: Path, emotions: tuple[str, str]) -> None:
    workbook = Workbook()
    index = workbook.active
    index.title = "index"
    index.append(["sheet_name", "chat_id"])
    index.append(["chat-a", "chat-a"])
    sheet = workbook.create_sheet("chat-a")
    sheet.append(
        [
            "轮次",
            "提问人员",
            "被谈话人",
            "预测情绪标签",
            "原始情绪标签",
            "预测原因",
            "原始模型输出",
        ]
    )
    for turn, emotion in enumerate(emotions, start=1):
        sheet.append(
            [
                turn,
                f"问题{turn}",
                f"回答{turn}",
                emotion,
                emotion,
                f"原因{turn}",
                f"原始输出{turn}",
            ]
        )
    workbook.save(path)


class MergeExperimentTests(unittest.TestCase):
    def test_merge_by_sheet_and_turn(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            file1 = root / "experiment1.xlsx"
            file2 = root / "experiment2.xlsx"
            output = root / "comparison.xlsx"
            create_experiment(file1, ("neutral", "fear"))
            create_experiment(file2, ("neutral", "nervousness"))

            report = merge_experiment_workbooks(file1, file2, output)

            self.assertEqual(1, report.compared_sheets)
            self.assertEqual(2, report.compared_turns)
            self.assertEqual(1, report.different_emotions)

            workbook = load_workbook(output)
            self.assertEqual(["index", "chat-a"], workbook.sheetnames)
            sheet = workbook["chat-a"]
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(
                ["轮次", "提问人员", "被谈话人", *COMPARISON_HEADERS], headers
            )
            self.assertEqual("neutral", sheet.cell(2, 4).value)
            self.assertEqual("neutral", sheet.cell(2, 5).value)
            self.assertEqual("fear", sheet.cell(3, 4).value)
            self.assertEqual("nervousness", sheet.cell(3, 5).value)
            self.assertEqual("00FFF2CC", sheet.cell(3, 4).fill.fgColor.rgb)
            self.assertEqual(2, workbook["index"].max_column)
            workbook.close()

            reopened = load_workbook(output, read_only=True)
            self.assertEqual(2, len(reopened.sheetnames))
            reopened.close()

    def test_missing_sheet_does_not_stop_merge(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            file1 = root / "experiment1.xlsx"
            file2 = root / "experiment2.xlsx"
            output = root / "comparison.xlsx"
            create_experiment(file1, ("neutral", "fear"))
            create_experiment(file2, ("neutral", "fear"))
            workbook = load_workbook(file2)
            del workbook["chat-a"]
            workbook.save(file2)

            report = merge_experiment_workbooks(file1, file2, output)
            self.assertEqual(["chat-a"], report.missing_sheets_in_file2)
            self.assertTrue(output.exists())


if __name__ == "__main__":
    unittest.main()
