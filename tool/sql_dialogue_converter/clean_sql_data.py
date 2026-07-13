"""Parse SQL dialogue records and create an Excel review workbook."""

from __future__ import annotations

import argparse
import json
import logging
import math
import re
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


LOGGER = logging.getLogger("sql_dialogue_parser")
ROLE_PATTERN = re.compile(r"(审调人员|被谈话人)\s*[:：]")
INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")
COMMON_PUNCTUATION = re.compile(r"[\s\u3000，。！？；：、,.!?;:'\"“”‘’（）()【】\[\]{}<>《》—_-]+")

AI_IDENTITY_KEYWORDS = (
    "语言模型",
    "通义千问",
    "作为AI",
    "作为 AI",
    "作为人工智能",
    "阿里云开发的一款",
    "我的目标是帮助用户",
    "我是一款大模型",
)
DEVICE_KEYWORDS = (
    "按v键",
    "按 V 键",
    "按V键",
    "点击按钮",
    "发送给模型",
    "打断操作",
    "麦克风",
    "设备调试",
    "点亮",
    "松手",
    "操作界面",
)

INDEX_HEADERS = (
    "sheet_name",
    "chat_id",
    "first_create_time",
    "last_create_time",
    "question_count",
    "answer_count",
    "complete_turn_count",
    "turn_count_mismatch",
    "dialogue_group",
    "ai_identity_contamination",
    "device_instruction_contamination",
    "high_repetition",
    "possible_transcription_issue",
    "review_priority",
    "contamination_flags",
)


@dataclass(frozen=True)
class SqlRecord:
    """One parsed SQL dialogue record."""

    chat_id: str
    message: str
    create_time: datetime
    record_number: int
    line_number: int


@dataclass(frozen=True)
class DialogueTurn:
    """One question/answer turn. Either side may be empty."""

    question: str = ""
    answer: str = ""


@dataclass
class ChatReview:
    """Aggregated dialogue and quality metadata for one chat ID."""

    chat_id: str
    records: list[SqlRecord]
    turns: list[DialogueTurn]
    sheet_name: str = ""
    flags: dict[str, bool] = field(default_factory=dict)

    @property
    def question_count(self) -> int:
        return sum(bool(turn.question) for turn in self.turns)

    @property
    def answer_count(self) -> int:
        return sum(bool(turn.answer) for turn in self.turns)

    @property
    def complete_turn_count(self) -> int:
        return sum(bool(turn.question and turn.answer) for turn in self.turns)

    @property
    def first_create_time(self) -> datetime:
        return min(record.create_time for record in self.records)

    @property
    def last_create_time(self) -> datetime:
        return max(record.create_time for record in self.records)


@dataclass(frozen=True)
class ParseResult:
    """SQL parser output with failures retained for reporting."""

    records: list[SqlRecord]
    total_records: int
    failures: list[dict[str, object]]


def parse_arguments() -> argparse.Namespace:
    """Parse command-line options."""
    parser = argparse.ArgumentParser(
        description="将 SQL 中的审讯对话整理为一个 Excel 复核工作簿"
    )
    parser.add_argument("input", type=Path, help="输入 SQL 文件")
    parser.add_argument(
        "--output-dir", type=Path, default=Path("output"), help="输出目录"
    )
    parser.add_argument(
        "--output-file", default="dialogue_review.xlsx", help="Excel 输出文件名"
    )
    return parser.parse_args()


def read_sql_file(path: Path) -> str:
    """Read a SQL file as UTF-8 and provide a clear encoding error."""
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError as exc:
        raise ValueError(f"SQL 文件不是有效的 UTF-8 编码：{path}") from exc


def _split_sql_statements(sql_text: str) -> list[tuple[str, int]]:
    """Split statements on semicolons outside SQL quoted strings."""
    statements: list[tuple[str, int]] = []
    start = 0
    start_line = 1
    line = 1
    in_quote = False
    index = 0

    while index < len(sql_text):
        char = sql_text[index]
        if char == "\n":
            line += 1
        if char == "'":
            if in_quote and index + 1 < len(sql_text) and sql_text[index + 1] == "'":
                index += 2
                continue
            if index == 0 or sql_text[index - 1] != "\\":
                in_quote = not in_quote
        elif char == ";" and not in_quote:
            statement = sql_text[start : index + 1]
            if statement.strip():
                statements.append((statement, start_line))
            start = index + 1
            start_line = line
        index += 1

    trailing = sql_text[start:]
    if trailing.strip():
        statements.append((trailing, start_line))
    return statements


def _parse_sql_tuple(tuple_text: str) -> list[str | None]:
    """Parse one VALUES tuple without splitting commas inside strings."""
    values: list[str | None] = []
    index = 1
    end = len(tuple_text) - 1

    while index < end:
        while index < end and (tuple_text[index].isspace() or tuple_text[index] == ","):
            index += 1
        if index >= end:
            break

        if tuple_text[index] == "'":
            index += 1
            chars: list[str] = []
            while index < end:
                char = tuple_text[index]
                if char == "'":
                    if index + 1 < end and tuple_text[index + 1] == "'":
                        chars.append("'")
                        index += 2
                        continue
                    index += 1
                    break
                if char == "\\" and index + 1 < end:
                    escaped = tuple_text[index + 1]
                    chars.append({"n": "\n", "r": "\r", "t": "\t"}.get(escaped, escaped))
                    index += 2
                    continue
                chars.append(char)
                index += 1
            else:
                raise ValueError("SQL 字符串缺少结束单引号")
            values.append("".join(chars))
        else:
            value_start = index
            while index < end and tuple_text[index] != ",":
                index += 1
            raw_value = tuple_text[value_start:index].strip()
            values.append(None if raw_value.upper() == "NULL" else raw_value)

    return values


def _extract_value_tuples(values_text: str) -> list[str]:
    """Extract parenthesized VALUES tuples while respecting quoted strings."""
    tuples: list[str] = []
    depth = 0
    start: int | None = None
    in_quote = False
    index = 0

    while index < len(values_text):
        char = values_text[index]
        if char == "'":
            if in_quote and index + 1 < len(values_text) and values_text[index + 1] == "'":
                index += 2
                continue
            if index == 0 or values_text[index - 1] != "\\":
                in_quote = not in_quote
        elif not in_quote:
            if char == "(":
                if depth == 0:
                    start = index
                depth += 1
            elif char == ")":
                depth -= 1
                if depth < 0:
                    raise ValueError("VALUES 括号不匹配")
                if depth == 0 and start is not None:
                    tuples.append(values_text[start : index + 1])
                    start = None
        index += 1

    if in_quote or depth != 0:
        raise ValueError("VALUES 中存在未闭合的引号或括号")
    return tuples


def _parse_create_time(value: str) -> datetime:
    """Parse common SQL datetime forms."""
    normalized = value.strip()
    try:
        return datetime.fromisoformat(normalized)
    except ValueError as exc:
        raise ValueError(f"无法解析 create_time：{value!r}") from exc


def parse_sql_records(sql_text: str) -> ParseResult:
    """Parse INSERT statements containing chat_id, message and create_time."""
    records: list[SqlRecord] = []
    failures: list[dict[str, object]] = []
    candidate_number = 0

    insert_pattern = re.compile(
        r"INSERT\s+INTO\s+.*?\((?P<columns>[^()]*)\)\s*VALUES\s*(?P<values>.*)",
        re.IGNORECASE | re.DOTALL,
    )

    for statement, line_number in _split_sql_statements(sql_text):
        if not re.search(r"\bINSERT\s+INTO\b", statement, re.IGNORECASE):
            continue
        candidate_number += 1
        try:
            match = insert_pattern.search(statement.rstrip().rstrip(";"))
            if not match:
                raise ValueError("无法识别 INSERT 列名或 VALUES")
            columns = [column.strip().strip("`").lower() for column in match.group("columns").split(",")]
            required = {"chat_id", "message", "create_time"}
            if not required.issubset(columns):
                raise ValueError(f"缺少必要列：{sorted(required - set(columns))}")

            tuples = _extract_value_tuples(match.group("values"))
            if not tuples:
                raise ValueError("VALUES 中没有数据元组")
            for tuple_text in tuples:
                values = _parse_sql_tuple(tuple_text)
                if len(values) != len(columns):
                    raise ValueError(f"列数为 {len(columns)}，但值数量为 {len(values)}")
                row = dict(zip(columns, values))
                if any(row[name] is None for name in required):
                    raise ValueError("必要字段不能为 NULL")
                records.append(
                    SqlRecord(
                        chat_id=str(row["chat_id"]),
                        message=str(row["message"]),
                        create_time=_parse_create_time(str(row["create_time"])),
                        record_number=candidate_number,
                        line_number=line_number,
                    )
                )
        except Exception as exc:  # One malformed record must not stop the file.
            failure = {
                "record_number": candidate_number,
                "line_number": line_number,
                "error": str(exc),
            }
            failures.append(failure)
            LOGGER.warning(
                "第 %s 条 INSERT（约第 %s 行）解析失败：%s",
                candidate_number,
                line_number,
                exc,
            )

    return ParseResult(records=records, total_records=candidate_number, failures=failures)


def _clean_role_content(content: str) -> str:
    """Remove separators around role blocks without altering inner punctuation."""
    return content.strip().strip(" \t\r\n;；")


def split_dialogue_turns(message: str) -> list[DialogueTurn]:
    """Split dialogue by role-marker positions, retaining incomplete turns."""
    markers = list(ROLE_PATTERN.finditer(message))
    if not markers:
        return []

    turns: list[DialogueTurn] = []
    pending_question: str | None = None
    for position, marker in enumerate(markers):
        content_end = markers[position + 1].start() if position + 1 < len(markers) else len(message)
        content = _clean_role_content(message[marker.end() : content_end])
        role = marker.group(1)

        if role == "审调人员":
            if pending_question is not None:
                turns.append(DialogueTurn(question=pending_question))
            pending_question = content
        elif pending_question is not None:
            turns.append(DialogueTurn(question=pending_question, answer=content))
            pending_question = None
        else:
            turns.append(DialogueTurn(answer=content))

    if pending_question is not None:
        turns.append(DialogueTurn(question=pending_question))
    return turns


def group_records_by_chat_id(records: Iterable[SqlRecord]) -> list[ChatReview]:
    """Sort records by time and aggregate their turns per chat ID."""
    grouped: dict[str, list[SqlRecord]] = defaultdict(list)
    for record in records:
        grouped[record.chat_id].append(record)

    reviews: list[ChatReview] = []
    for chat_id, chat_records in grouped.items():
        chat_records.sort(key=lambda record: (record.create_time, record.record_number))
        turns = [
            turn
            for record in chat_records
            for turn in split_dialogue_turns(record.message)
        ]
        reviews.append(ChatReview(chat_id=chat_id, records=chat_records, turns=turns))
    return reviews


def _normalize_question(text: str) -> str:
    return COMMON_PUNCTUATION.sub("", text).lower()


def calculate_quality_flags(review: ChatReview) -> dict[str, bool]:
    """Calculate conservative, explainable contamination and quality flags."""
    combined_text = "\n".join(
        part for turn in review.turns for part in (turn.question, turn.answer) if part
    )
    normalized_questions = [
        normalized
        for turn in review.turns
        if (normalized := _normalize_question(turn.question))
    ]
    repetitions = Counter(normalized_questions)
    question_count = review.question_count
    answer_count = review.answer_count
    difference = abs(question_count - answer_count)
    severe_mismatch = difference >= max(2, math.ceil(max(question_count, answer_count, 1) * 0.5))
    suspicious_repeat = bool(re.search(r"([\u4e00-\u9fff])\1(?=[\u4e00-\u9fff])", combined_text))

    return {
        "ai_identity_contamination": any(keyword in combined_text for keyword in AI_IDENTITY_KEYWORDS),
        "device_instruction_contamination": any(keyword in combined_text for keyword in DEVICE_KEYWORDS),
        "high_repetition": any(count >= 3 for count in repetitions.values()),
        "possible_transcription_issue": severe_mismatch or suspicious_repeat,
    }


def calculate_dialogue_group(complete_turn_count: int) -> str:
    """Assign a length group from complete question/answer turns."""
    if complete_turn_count >= 15:
        return "long"
    if complete_turn_count >= 8:
        return "medium"
    if complete_turn_count >= 4:
        return "short"
    return "very_short"


def calculate_review_priority(review: ChatReview) -> str:
    """Assign review priority using isolated, easily adjustable rules."""
    flags = review.flags
    if (
        review.complete_turn_count <= 3
        or flags["ai_identity_contamination"]
        or flags["device_instruction_contamination"]
    ):
        return "low"
    if review.complete_turn_count >= 8 and not any(flags.values()):
        return "high"
    return "medium"


def sanitize_sheet_name(raw_name: str, used_names: set[str]) -> str:
    """Return a legal, case-insensitively unique Excel sheet name."""
    base = INVALID_SHEET_CHARS.sub("_", raw_name).strip().strip("'") or "chat"
    base = base[:31]
    candidate = base
    suffix_number = 2
    normalized_used = {name.casefold() for name in used_names}
    while candidate.casefold() in normalized_used:
        suffix = f"_{suffix_number}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        suffix_number += 1
    used_names.add(candidate)
    return candidate


def _apply_table_style(
    worksheet: Worksheet,
    widths: Sequence[float],
    max_row_height: float = 120,
) -> None:
    """Apply shared readable Excel table formatting."""
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    for row_index in range(2, worksheet.max_row + 1):
        text_lengths = [len(str(worksheet.cell(row_index, column).value or "")) for column in range(1, worksheet.max_column + 1)]
        estimated_lines = max(
            1,
            max(math.ceil(length / max(widths[index] - 2, 1)) for index, length in enumerate(text_lengths)),
        )
        worksheet.row_dimensions[row_index].height = min(max_row_height, 18 * estimated_lines)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _flags_as_text(flags: dict[str, bool]) -> str:
    labels = {
        "ai_identity_contamination": "AI身份污染",
        "device_instruction_contamination": "设备操作说明污染",
        "high_repetition": "高重复内容",
        "possible_transcription_issue": "疑似转写异常",
    }
    return "；".join(labels[key] for key, value in flags.items() if value)


def create_index_sheet(workbook: Workbook, reviews: Sequence[ChatReview]) -> Worksheet:
    """Create the first-sheet chat index."""
    worksheet = workbook.active
    worksheet.title = "index"
    worksheet.append(INDEX_HEADERS)
    for review in reviews:
        worksheet.append(
            (
                review.sheet_name,
                review.chat_id,
                review.first_create_time,
                review.last_create_time,
                review.question_count,
                review.answer_count,
                review.complete_turn_count,
                review.question_count != review.answer_count,
                calculate_dialogue_group(review.complete_turn_count),
                review.flags["ai_identity_contamination"],
                review.flags["device_instruction_contamination"],
                review.flags["high_repetition"],
                review.flags["possible_transcription_issue"],
                calculate_review_priority(review),
                _flags_as_text(review.flags),
            )
        )
    _apply_table_style(worksheet, [24, 24, 20, 20, 15, 15, 20, 20, 16, 24, 30, 18, 28, 18, 40])
    return worksheet


def create_dialogue_sheet(workbook: Workbook, review: ChatReview) -> Worksheet:
    """Create one three-column dialogue worksheet."""
    worksheet = workbook.create_sheet(review.sheet_name)
    worksheet.append(("轮次", "提问人员", "被谈话人"))
    for turn_number, turn in enumerate(review.turns, start=1):
        worksheet.append((turn_number, turn.question, turn.answer))
    _apply_table_style(worksheet, [9, 70, 70])
    return worksheet


def prepare_reviews(records: Iterable[SqlRecord]) -> list[ChatReview]:
    """Aggregate, flag, sort and assign worksheet names."""
    reviews = group_records_by_chat_id(records)
    for review in reviews:
        review.flags = calculate_quality_flags(review)
    reviews.sort(key=lambda review: (-review.complete_turn_count, review.first_create_time, review.chat_id))

    used_names = {"index"}
    for review in reviews:
        review.sheet_name = sanitize_sheet_name(review.chat_id, used_names)
    return reviews


def write_excel_workbook(reviews: Sequence[ChatReview], output_path: Path) -> None:
    """Write the index and all chat worksheets to an XLSX file."""
    workbook = Workbook()
    create_index_sheet(workbook, reviews)
    for review in reviews:
        create_dialogue_sheet(workbook, review)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_summary(parse_result: ParseResult, reviews: Sequence[ChatReview]) -> dict[str, object]:
    """Build the machine-readable screening summary."""
    turn_counts = [len(review.turns) for review in reviews]
    group_counts = Counter(calculate_dialogue_group(review.complete_turn_count) for review in reviews)
    priority_counts = Counter(calculate_review_priority(review) for review in reviews)
    contamination_counts = {
        key: sum(review.flags[key] for review in reviews)
        for key in (
            "ai_identity_contamination",
            "device_instruction_contamination",
            "high_repetition",
            "possible_transcription_issue",
        )
    }
    return {
        "total_records": parse_result.total_records,
        "total_chat_ids": len(reviews),
        "total_turns": sum(turn_counts),
        "parse_failed_records": len(parse_result.failures),
        "failed_records": parse_result.failures,
        "group_counts": dict(group_counts),
        "priority_counts": dict(priority_counts),
        "contamination_counts": contamination_counts,
        "turn_statistics": {
            "min": min(turn_counts, default=0),
            "max": max(turn_counts, default=0),
            "mean": statistics.mean(turn_counts) if turn_counts else 0,
            "median": statistics.median(turn_counts) if turn_counts else 0,
        },
    }


def write_summary_json(summary: dict[str, object], output_path: Path) -> None:
    """Write a UTF-8 JSON summary."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def run(input_path: Path, output_dir: Path, output_file: str) -> tuple[Path, Path, dict[str, object]]:
    """Execute the complete SQL-to-review-workbook pipeline."""
    if not input_path.is_file():
        raise FileNotFoundError(f"找不到输入 SQL 文件：{input_path}")
    if Path(output_file).name != output_file:
        raise ValueError("--output-file 只能是文件名；输出目录请使用 --output-dir")
    if not output_file.lower().endswith(".xlsx"):
        raise ValueError("--output-file 必须使用 .xlsx 扩展名")

    LOGGER.info("正在读取 SQL 文件：%s", input_path)
    parse_result = parse_sql_records(read_sql_file(input_path))
    reviews = prepare_reviews(parse_result.records)
    excel_path = output_dir / output_file
    summary_path = output_dir / "dialogue_screening_summary.json"

    write_excel_workbook(reviews, excel_path)
    summary = build_summary(parse_result, reviews)
    write_summary_json(summary, summary_path)
    LOGGER.info(
        "处理完成：%s 条记录，%s 个 chat_id，%s 轮对话，%s 条失败",
        parse_result.total_records,
        len(reviews),
        summary["total_turns"],
        len(parse_result.failures),
    )
    LOGGER.info("Excel：%s", excel_path.resolve())
    LOGGER.info("JSON：%s", summary_path.resolve())
    return excel_path, summary_path, summary


def main() -> None:
    """CLI entry point."""
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    args = parse_arguments()
    run(args.input, args.output_dir, args.output_file)


if __name__ == "__main__":
    main()
