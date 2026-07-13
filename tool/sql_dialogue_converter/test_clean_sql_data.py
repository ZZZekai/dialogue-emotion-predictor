"""Tests for the SQL dialogue review tool."""

from __future__ import annotations

import json
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

from clean_sql_data import (
    ChatReview,
    DialogueTurn,
    SqlRecord,
    calculate_quality_flags,
    parse_sql_records,
    prepare_reviews,
    run,
    sanitize_sheet_name,
    split_dialogue_turns,
)


def insert_sql(chat_id: str, message: str, create_time: str) -> str:
    escaped = message.replace("'", "''")
    return (
        "INSERT INTO `dialogue` (`chat_id`, `message`, `create_time`) VALUES "
        f"('{chat_id}', '{escaped}', '{create_time}');"
    )


class DialogueSplitTests(unittest.TestCase):
    def test_multiple_turns_and_inner_punctuation(self) -> None:
        message = "审调人员：问题一，含逗号；继续？\n被谈话人:回答一；仍是回答。审调人员:问题二;被谈话人：回答二"
        turns = split_dialogue_turns(message)
        self.assertEqual(2, len(turns))
        self.assertIn("含逗号；继续", turns[0].question)
        self.assertEqual("回答二", turns[1].answer)

    def test_question_without_answer(self) -> None:
        self.assertEqual([DialogueTurn(question="问题")], split_dialogue_turns("审调人员：问题"))

    def test_answer_without_leading_question(self) -> None:
        turns = split_dialogue_turns("被谈话人：先回答。审调人员：再问")
        self.assertEqual("先回答。", turns[0].answer)
        self.assertEqual("再问", turns[1].question)


class SqlParserTests(unittest.TestCase):
    def test_escaped_quote_and_multiple_chat_ids(self) -> None:
        sql = "\n".join(
            (
                insert_sql("chat-1", "审调人员：你的名字？被谈话人：我叫 O'Brien。", "2026-01-01 10:00:00"),
                insert_sql("chat-2", "审调人员:问题被谈话人:回答", "2026-01-02 10:00:00"),
            )
        )
        result = parse_sql_records(sql)
        self.assertEqual(2, len(result.records))
        self.assertIn("O'Brien", result.records[0].message)

    def test_malformed_insert_does_not_stop_parser(self) -> None:
        sql = "INSERT INTO broken VALUES ('x');\n" + insert_sql(
            "ok", "审调人员：问被谈话人：答", "2026-01-01 00:00:00"
        )
        result = parse_sql_records(sql)
        self.assertEqual(1, len(result.records))
        self.assertEqual(1, len(result.failures))

    def test_same_chat_records_are_sorted_and_merged(self) -> None:
        sql = "\n".join(
            (
                insert_sql("same", "审调人员：后问被谈话人：后答", "2026-01-02 00:00:00"),
                insert_sql("same", "审调人员：先问被谈话人：先答", "2026-01-01 00:00:00"),
            )
        )
        reviews = prepare_reviews(parse_sql_records(sql).records)
        self.assertEqual(["先问", "后问"], [turn.question for turn in reviews[0].turns])


class QualityAndExcelTests(unittest.TestCase):
    def _review(self, turns: list[DialogueTurn]) -> ChatReview:
        record = SqlRecord("id", "", datetime(2026, 1, 1), 1, 1)
        return ChatReview("id", [record], turns)

    def test_contamination_and_repetition_flags(self) -> None:
        review = self._review(
            [
                DialogueTurn("你叫什么名字？", "作为AI语言模型，我按v键操作。"),
                DialogueTurn("你叫什么名字？", "回答"),
                DialogueTurn("你叫什么名字？", "回答"),
            ]
        )
        flags = calculate_quality_flags(review)
        self.assertTrue(flags["ai_identity_contamination"])
        self.assertTrue(flags["device_instruction_contamination"])
        self.assertTrue(flags["high_repetition"])

    def test_sheet_name_is_legal_and_unique(self) -> None:
        used = {"index"}
        first = sanitize_sheet_name("bad/name:*?[]", used)
        second = sanitize_sheet_name("bad/name:*?[]", used)
        self.assertNotEqual(first, second)
        self.assertNotRegex(first, r"[:\\/?*\[\]]")

    def test_end_to_end_workbook_can_be_reopened(self) -> None:
        sql = "\n".join(
            (
                insert_sql("chat/one", "审调人员：问题1被谈话人：回答1", "2026-01-02 00:00:00"),
                insert_sql("chat/one", "审调人员：问题2", "2026-01-03 00:00:00"),
                insert_sql("chat-two", "被谈话人：开头回答", "2026-01-01 00:00:00"),
            )
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_path = root / "input.sql"
            input_path.write_text(sql, encoding="utf-8")
            excel_path, summary_path, summary = run(input_path, root, "review.xlsx")

            workbook = load_workbook(excel_path)
            self.assertEqual("index", workbook.sheetnames[0])
            self.assertEqual(3, len(workbook.sheetnames))
            self.assertEqual(["轮次", "提问人员", "被谈话人"], [cell.value for cell in workbook[workbook.sheetnames[1]][1]])
            self.assertEqual(2, summary["total_chat_ids"])
            self.assertEqual(3, summary["total_turns"])
            self.assertEqual(summary, json.loads(summary_path.read_text(encoding="utf-8")))


if __name__ == "__main__":
    unittest.main()
